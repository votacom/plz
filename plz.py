#!/usr/bin/python

"""
Append to a spreadsheet with a postal code column columns for the geographic coordinates of the postal code.
"""

import requests
import json
import argparse
import os.path
import sys
import openpyxl
from openpyxl import Workbook

def get_plzs(overpass_url, country="AT"):
    """Featch via Overpass API all postal codes in the country with the given ISO3166-1 code. Return a JSON with those postal codes together with their geographic coordinates."""
    # query to get all postal codes in austria together with the geographic center of their area:
    overpass_query = """
    [out:json];
    area["ISO3166-1"={}][admin_level=2];
    relation[boundary=postal_code](area);
    out tags center;
    """.format(country)
    response = requests.get(overpass_url, params={'data': overpass_query})
    data = response.json()
    return data

def initgeo(overpass_url="http://overpass-api.de/api/interpreter", plz_file="plz.json"):
    """If PLZ file exists, read data from there. Otherwise, consult overpass API to fetch the PLZ data and save it to PLZ file. Return a dict of plz-(lat,lon) pairs."""
    if not os.path.exists( plz_file ):
        print( 'PLZ file {} not found. Fetching PLZ data from {}...'.format( plz_file, overpass_url ) )
        data = get_plzs( overpass_url )
        print( 'Saving fetched PLZ data in file {}.'.format( plz_file ) )
        with open( plz_file, 'w' ) as f:
            json.dump( data, f )
    else:
        f = open( plz_file )
        data = json.load( f )
        f.close()
    # now we have the geographic PLZ data in the JSON object `data`. Let's convert that into an easy-to-use key-value-map (key=postcode, value=pair of lat+lon):
    return {element['tags']['postal_code'] : element['center'] for element in data['elements']}

def main():
    # parse command-line arguments:
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument('spreadsheet', help='file name of the spreadsheet in which to append coordinate columns')
    parser.add_argument('-c', '--plzcolumn', help='header name of the spreadsheet column holding the PLZ number', default='PLZ')
    parser.add_argument('-j', '--plzfile', help='file name of the JSON-formatted PLZ-coordinate map', default='plz.json')
    parser.add_argument('-u', '--overpassurl', help='URL of the Overpass API to use to fetch geographic data', default='http://overpass-api.de/api/interpreter')
    args = parser.parse_args()

    geo = initgeo( args.overpassurl, args.plzfile )

    # Proceed with opening the spreadsheet:
    workbook = openpyxl.load_workbook( filename=args.spreadsheet )
    sheet = workbook.active
    print( 'working in active workbook sheet {}'.format( sheet.title ) )
    header_row = sheet[1]
    plz_columns = [ cell for cell in header_row if cell.value == args.plzcolumn ]
    if len(plz_columns) == 0:
        sys.exit( 'spreadsheet has no heading cell in the first row containing "{}"'.format( args.plzcolumn ) )
    plz_col = plz_columns[0].column # attention: 1-based!
    print( 'identified column {} as the input PLZ column'.format( plz_col ) )
    # let's find out if there are already longitude and latutude columns and if not, inserte them anew after the PLZ column:
    lat_columns = [ cell for cell in header_row if cell.value == 'Latitude' ]
    lon_columns = [ cell for cell in header_row if cell.value == 'Longitude' ]
    if len(lat_columns) == 0:
        # create Latitude column:
        lat_col = plz_col + 1
        sheet.insert_cols( lat_col )
        sheet.cell(row=1, column=lat_col).value = 'Latitude'
    else:
        lat_col = lat_columns[0].column
    if len(lon_columns) == 0:
        # create Longitude column:
        lon_col = lat_col + 1
        sheet.insert_cols( lon_col )
        sheet.cell(row=1, column=lon_col).value = 'Longitude'
    else:
        lon_col = lon_columns[0].column
    # now, let's iterate over the data records and lookup their PLZ in ´geo´:
    for row in sheet.iter_rows( min_row=2 ):
        plz = row[plz_col-1].value #-1 to get 0-based.
        plz = str(plz) # if encoded as Number in spreadsheet, transform to string to map the dict
        if plz in geo:
            row[lat_col-1].value = geo[plz]['lat']
            row[lon_col-1].value = geo[plz]['lon']
        else:
            print( 'PLZ {} not found in PLZ database. Skipping row {}'.format( plz, row[0].row ) )
    # finished. save and exit:
    workbook.save( args.spreadsheet )

if __name__ == '__main__':
    main()
